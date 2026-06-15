#!/usr/bin/env python3
"""Validate Korea SW project FP workbook sheets.

Korean: 대한민국 SW사업 대가산정 가이드 기준으로 FP산정 시트의 기능유형, 복잡도,
가중치, 중복 후보, 비용시트 합계를 점검한다.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

VALID_TYPES = {"ILF", "EIF", "EI", "EO", "EQ"}
TRADITIONAL_WEIGHTS = {
    "ILF": {"L": Decimal("7"), "A": Decimal("10"), "H": Decimal("15")},
    "EIF": {"L": Decimal("5"), "A": Decimal("7"), "H": Decimal("10")},
    "EI": {"L": Decimal("3"), "A": Decimal("4"), "H": Decimal("6")},
    "EO": {"L": Decimal("4"), "A": Decimal("5"), "H": Decimal("7")},
    "EQ": {"L": Decimal("3"), "A": Decimal("4"), "H": Decimal("6")},
}
SIMPLIFIED_WEIGHTS = {
    "ILF": Decimal("7.5"),
    "EIF": Decimal("5.4"),
    "EI": Decimal("4.0"),
    "EO": Decimal("5.2"),
    "EQ": Decimal("3.9"),
}


@dataclass
class RowResult:
    status: str
    excel_row: int
    application: str | None
    detail_work: str | None
    unit_process: str | None
    description: str | None
    fp_type: str | None
    ret_ftr: Any
    det: Any
    submitted_complexity: Any
    expected_complexity: str | None
    submitted_weight: Any
    expected_weight: Decimal | None
    note: str | None


@dataclass
class Finding:
    severity: str
    category: str
    excel_row: int | None
    field: str
    fp_type: str | None
    application: str | None
    detail_work: str | None
    unit_process: str | None
    submitted: Any
    expected: Any
    reason_ko: str
    recommendation_ko: str


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate Korea FP workbook")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default=None, help="FP sheet name. Auto-detects FP산정/FP산정(간이법).")
    parser.add_argument("--method", choices=["traditional", "simplified", "auto"], default="auto")
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--report-name", default=None)
    args = parser.parse_args()

    source = args.workbook.expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(source)

    output_dir = args.output_dir or source.parent / "fp_validation_report"
    output_dir.mkdir(parents=True, exist_ok=True)
    report_name = args.report_name or safe_stem(source.stem) + "_fp_validation"

    rows, non_data_rows, summary = validate_workbook(source, args.sheet, args.method)
    findings = build_findings(rows)
    add_duplicates(rows, findings)
    add_review_patterns(rows, findings)
    apply_status(rows, findings)
    summary.update(build_summary(source, rows, non_data_rows, findings))

    xlsx_path = output_dir / f"{report_name}.xlsx"
    json_path = output_dir / f"{report_name}.json"
    md_path = output_dir / f"{report_name}.md"
    write_xlsx(xlsx_path, rows, non_data_rows, findings, summary)
    json_path.write_text(json.dumps({"summary": summary, "findings": [asdict(f) for f in findings]}, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    write_md(md_path, summary)

    print(json.dumps({"xlsx": str(xlsx_path), "json": str(json_path), "md": str(md_path), "summary": summary}, ensure_ascii=False, indent=2, default=str))


def validate_workbook(source: Path, sheet_name: str | None, method: str) -> tuple[list[RowResult], list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(source, read_only=True, data_only=True, keep_vba=False)
    sheet = sheet_name or detect_fp_sheet(wb.sheetnames)
    if sheet is None:
        raise ValueError(f"Could not detect FP sheet from {wb.sheetnames}")
    inferred_method = method if method != "auto" else ("simplified" if "간이" in sheet else "traditional")
    ws = wb[sheet]
    header_row = detect_header_row(ws, inferred_method)
    data_start = header_row + 1
    rows: list[RowResult] = []
    non_data_rows: list[dict[str, Any]] = []

    if inferred_method == "traditional":
        min_col, max_col = 2, 11
    else:
        min_col, max_col = 2, 8

    for excel_row, values_tuple in enumerate(ws.iter_rows(min_row=data_start, min_col=min_col, max_col=max_col, values_only=True), start=data_start):
        values = list(values_tuple)
        if blank_row(values):
            continue
        if inferred_method == "traditional":
            row = RowResult("OK", excel_row, clean(values[0]), clean(values[1]), clean(values[2]), clean(values[3]), upper(values[4]), values[5], values[6], values[7], None, values[8], None, clean(values[9]))
        else:
            row = RowResult("OK", excel_row, clean(values[0]), clean(values[1]), clean(values[2]), clean(values[3]), upper(values[4]), None, None, None, None, values[5], None, clean(values[6]) if len(values) > 6 else None)
        if row.fp_type in VALID_TYPES:
            if inferred_method == "traditional":
                ret_ftr = positive_int(row.ret_ftr)
                det = positive_int(row.det)
                if ret_ftr is not None and det is not None:
                    row.expected_complexity = traditional_complexity(row.fp_type, ret_ftr, det)
                    row.expected_weight = TRADITIONAL_WEIGHTS[row.fp_type][row.expected_complexity]
            else:
                row.expected_weight = SIMPLIFIED_WEIGHTS[row.fp_type]
            rows.append(row)
        else:
            non_data_rows.append({"excel_row": excel_row, "values": values, "reason_ko": "FP유형이 비어 있거나 ILF/EIF/EI/EO/EQ가 아니어서 데이터 행에서 제외됨"})

    return rows, non_data_rows, {"source_file": unicodedata.normalize("NFC", source.name), "fp_sheet": sheet, "method": inferred_method, **read_cost_summary(wb)}


def build_findings(rows: list[RowResult]) -> list[Finding]:
    findings: list[Finding] = []
    for row in rows:
        for field, value, label in [("application", row.application, "어플리케이션명"), ("detail_work", row.detail_work, "세부 업무명"), ("unit_process", row.unit_process, "단위프로세스명"), ("fp_type", row.fp_type, "FP유형")]:
            if not value:
                findings.append(finding("ERROR", "필수값", row, field, value, "필수 입력", f"{label}이 비어 있습니다.", f"{label}을 입력하십시오."))
        if row.fp_type not in VALID_TYPES:
            findings.append(finding("ERROR", "FP유형", row, "fp_type", row.fp_type, "ILF/EIF/EI/EO/EQ", "FP유형이 가이드 기능유형이 아닙니다.", "유효한 FP유형으로 정리하십시오."))
            continue
        if row.expected_weight is None:
            findings.append(finding("ERROR", "산정입력", row, "ret_ftr/det", f"{row.ret_ftr}/{row.det}", "양의 정수", "RET/FTR 또는 DET가 비어 있거나 산정 불가합니다.", "가이드 기준으로 RET/FTR 및 DET를 입력하십시오."))
            continue
        submitted_weight = decimal_or_none(row.submitted_weight)
        if row.expected_complexity and clean(row.submitted_complexity) and clean(row.submitted_complexity).upper() != row.expected_complexity:
            findings.append(finding("ERROR", "복잡도", row, "complexity", row.submitted_complexity, row.expected_complexity, "복잡도가 가이드 매트릭스 재계산 결과와 다릅니다.", "FPR/복잡도 값을 재계산하십시오."))
        if submitted_weight is None:
            findings.append(finding("ERROR", "가중치", row, "weight", row.submitted_weight, row.expected_weight, "가중치가 비어 있거나 숫자가 아닙니다.", "가이드 기준 가중치를 입력하십시오."))
        elif submitted_weight != row.expected_weight:
            findings.append(finding("ERROR", "가중치", row, "weight", submitted_weight, row.expected_weight, "가중치가 가이드 기준과 다릅니다.", "FP유형/복잡도별 가중치를 재확인하십시오."))
    return findings


def add_duplicates(rows: list[RowResult], findings: list[Finding]) -> None:
    seen: dict[tuple[str, str, str, str], RowResult] = {}
    for row in rows:
        if not row.unit_process:
            continue
        key = (norm(row.application), norm(row.detail_work), norm(row.unit_process), row.fp_type or "")
        if key in seen:
            first = seen[key]
            findings.append(finding("WARNING", "중복", row, "unit_process", row.unit_process, f"최초 행 {first.excel_row}", "동일 계층의 단위프로세스/FP유형 중복 후보입니다.", "별도 기본 프로세스인지 확인하고 중복이면 한 건만 남기십시오."))
        else:
            seen[key] = row


def add_review_patterns(rows: list[RowResult], findings: list[Finding]) -> None:
    patterns = ["이력", "로그", "임시", "백업", "첨부파일", "인터페이스", "코드"]
    for row in rows:
        text = " ".join(x for x in [row.unit_process, row.description] if x)
        if row.fp_type == "EIF":
            findings.append(finding("INFO", "EIF검토", row, "fp_type", row.fp_type, "외부 유지관리 논리파일", "EIF는 타 시스템이 유지관리하고 본 시스템이 참조하는 데이터여야 합니다.", "동일 사업 범위 내부 데이터라면 ILF로 재검토하십시오."))
        if row.fp_type in {"ILF", "EIF"} and any(p in text for p in patterns):
            findings.append(finding("INFO", "논리파일검토", row, "unit_process/description", text[:180], "독립 논리파일", "이력/로그/임시/첨부/코드성 데이터는 독립 ILF/EIF 여부 검토가 필요합니다.", "상위 업무 ILF의 하위 RET/DET인지 확인하십시오."))


def apply_status(rows: list[RowResult], findings: list[Finding]) -> None:
    by_row: dict[int, list[str]] = defaultdict(list)
    for item in findings:
        if item.excel_row:
            by_row[item.excel_row].append(item.severity)
    for row in rows:
        severities = by_row.get(row.excel_row, [])
        row.status = "ERROR" if "ERROR" in severities else "WARNING" if "WARNING" in severities else "INFO" if "INFO" in severities else "OK"


def build_summary(source: Path, rows: list[RowResult], non_data_rows: list[dict[str, Any]], findings: list[Finding]) -> dict[str, Any]:
    counts = Counter(r.fp_type for r in rows)
    submitted_fp = defaultdict(Decimal)
    expected_fp = defaultdict(Decimal)
    for row in rows:
        submitted = decimal_or_none(row.submitted_weight)
        if submitted is not None:
            submitted_fp[row.fp_type] += submitted
        if row.expected_weight is not None:
            expected_fp[row.fp_type] += row.expected_weight
    return {
        "data_rows": len(rows),
        "non_data_rows": len(non_data_rows),
        "function_count_by_type": dict(counts),
        "submitted_fp_by_type": {k: float(v) for k, v in submitted_fp.items()},
        "expected_fp_by_type": {k: float(v) for k, v in expected_fp.items()},
        "submitted_fp_total": float(sum(submitted_fp.values(), Decimal("0"))),
        "expected_fp_total": float(sum(expected_fp.values(), Decimal("0"))),
        "status_counts": dict(Counter(r.status for r in rows)),
        "finding_counts_by_severity": dict(Counter(f.severity for f in findings)),
        "finding_counts_by_category": dict(Counter(f.category for f in findings)),
    }


def write_xlsx(path: Path, rows: list[RowResult], non_data_rows: list[dict[str, Any]], findings: list[Finding], summary: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"
    write_table(ws, ["항목", "값"], [[k, json.dumps(v, ensure_ascii=False, default=str) if isinstance(v, (dict, list)) else v] for k, v in summary.items()])
    ws = wb.create_sheet("점검결과")
    write_table(ws, ["심각도", "분류", "원본행", "필드", "FP유형", "어플리케이션명", "세부 업무명", "단위프로세스명", "입력값", "기대값", "검토사유", "권고사항"], [[f.severity, f.category, f.excel_row, f.field, f.fp_type, f.application, f.detail_work, f.unit_process, str(f.submitted) if f.submitted is not None else None, str(f.expected) if f.expected is not None else None, f.reason_ko, f.recommendation_ko] for f in findings])
    ws = wb.create_sheet("FP행별점검")
    write_table(ws, ["상태", "원본행", "어플리케이션명", "세부 업무명", "단위프로세스명", "설명", "FP유형", "RET/FTR", "DET", "시트복잡도", "재계산복잡도", "시트가중치", "재계산가중치", "비고"], [[r.status, r.excel_row, r.application, r.detail_work, r.unit_process, r.description, r.fp_type, r.ret_ftr, r.det, r.submitted_complexity, r.expected_complexity, r.submitted_weight, r.expected_weight, r.note] for r in rows])
    ws = wb.create_sheet("비데이터행")
    write_table(ws, ["원본행", "값", "제외사유"], [[r["excel_row"], str(r["values"]), r["reason_ko"]] for r in non_data_rows])
    for sheet in wb.worksheets:
        style(sheet)
    wb.save(path)


def write_md(path: Path, summary: dict[str, Any]) -> None:
    lines = ["# Korea FP Workbook Validation", "", f"- Source: `{summary.get('source_file')}`", f"- FP sheet: `{summary.get('fp_sheet')}`", f"- Method: `{summary.get('method')}`", f"- Data rows: {summary.get('data_rows')}", f"- Submitted FP total: {summary.get('submitted_fp_total')}", f"- Expected FP total: {summary.get('expected_fp_total')}", "", "## Findings", ""]
    for k, v in (summary.get("finding_counts_by_severity") or {}).items():
        lines.append(f"- {k}: {v}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_table(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def style(ws: Any) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        status = row[0].value
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if status == "ERROR":
                cell.fill = PatternFill("solid", fgColor="F4CCCC")
            elif status == "WARNING":
                cell.fill = PatternFill("solid", fgColor="FCE5CD")
            elif status == "INFO":
                cell.fill = PatternFill("solid", fgColor="D9EAD3")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18 if idx < 10 else 48


def detect_fp_sheet(sheetnames: list[str]) -> str | None:
    for candidate in ["FP산정", "FP산정(간이법)"]:
        if candidate in sheetnames:
            return candidate
    for name in sheetnames:
        if "FP산정" in name:
            return name
    return None


def detect_header_row(ws: Any, method: str) -> int:
    required = ["FP유형"]
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True):
        row_num = row[0] if False else None
    for r in range(1, min(ws.max_row, 30) + 1):
        values = [clean(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 15) + 1)]
        if any(v and "FP유형" in v for v in values):
            return r
    return 5


def read_cost_summary(wb: Any) -> dict[str, Any]:
    for name in ["SW개발비 산정", "SW개발비산정", "SW유지관리비", "SW재개발비 산정"]:
        if name in wb.sheetnames:
            ws = wb[name]
            return {"cost_sheet": name, "cost_total_fp_candidate_B7": ws["B7"].value, "cost_unit_price_candidate_D7": ws["D7"].value, "cost_profit_rate_candidate_J9": ws["J9"].value}
    return {}


def traditional_complexity(fp_type: str, ret_ftr: int, det: int) -> str:
    if fp_type in {"ILF", "EIF"}:
        ret_band = 0 if ret_ftr <= 1 else 1 if ret_ftr <= 5 else 2
        det_band = band(det, 19, 50)
    elif fp_type == "EI":
        ret_band = 0 if ret_ftr <= 1 else 1 if ret_ftr == 2 else 2
        det_band = band(det, 4, 15)
    elif fp_type == "EO":
        ret_band = 0 if ret_ftr <= 1 else 1 if ret_ftr <= 3 else 2
        det_band = band(det, 5, 19)
    elif fp_type == "EQ":
        ret_band = 0 if ret_ftr == 1 else 1 if ret_ftr <= 3 else 2
        det_band = band(det, 5, 19)
    else:
        raise ValueError(fp_type)
    matrix = (("L", "L", "A"), ("L", "A", "H"), ("A", "H", "H"))
    return matrix[ret_band][det_band]


def band(value: int, low_max: int, average_max: int) -> int:
    return 0 if value <= low_max else 1 if value <= average_max else 2


def finding(severity: str, category: str, row: RowResult, field: str, submitted: Any, expected: Any, reason_ko: str, recommendation_ko: str) -> Finding:
    return Finding(severity, category, row.excel_row, field, row.fp_type, row.application, row.detail_work, row.unit_process, submitted, expected, reason_ko, recommendation_ko)


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def upper(value: Any) -> str | None:
    text = clean(value)
    return text.upper() if text else None


def blank_row(values: list[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def positive_int(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        dec = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        return None
    if dec <= 0 or dec != dec.to_integral_value():
        return None
    return int(dec)


def decimal_or_none(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool) or value == "":
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        return None


def norm(value: str | None) -> str:
    return re.sub(r"[^0-9a-z가-힣]+", "", (value or "").casefold())


def safe_stem(value: str) -> str:
    value = unicodedata.normalize("NFC", value)
    return re.sub(r"[^0-9A-Za-z가-힣_.-]+", "_", value).strip("_") or "fp_workbook"


if __name__ == "__main__":
    main()
